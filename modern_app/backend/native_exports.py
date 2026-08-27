from __future__ import annotations

import csv
import base64
import html
import json
from collections import Counter
from datetime import datetime, timedelta
from pathlib import Path
from typing import Any

from database import CASE_COLUMNS, IN_PROGRESS_COLUMNS, connect, ensure_schema, get_json_setting
from paths import automated_reports_dir, logo_path, marker_icon_path


GRAPH_COLUMN_MAP = {
    "Offense Type": "offense_type",
    "Device Type": "device_type",
    "Agency": "agency",
    "Investigator": "investigator",
    "Forensic Tool": "forensic_tool",
    "City of Offense": "city_of_offense",
    "State of Offense": "state_of_offense",
    "Total Volume by Examiner": "examiner",
    "Total Volume by Investigator": "investigator",
    "Total Volume by Agency": "agency",
    "Total Volume by Device Type": "device_type",
}

VOLUME_GRAPH_TYPES = {
    "Total Volume by Examiner",
    "Total Volume by Investigator",
    "Total Volume by Agency",
    "Total Volume by Device Type",
}


def run_native_exports(config: dict[str, Any]) -> dict[str, Any]:
    ensure_schema()
    output_dir = _resolve_output_dir(config.get("output_dir"))
    output_dir.mkdir(parents=True, exist_ok=True)
    report_types = set(config.get("report_types") or [])

    completed_rows = _read_rows("case_log", CASE_COLUMNS)
    progress_rows = _read_rows("in_progress_cases", IN_PROGRESS_COLUMNS)
    scoped_rows = _scope_rows(completed_rows, config)

    files: list[dict[str, Any]] = []
    data_exports = config.get("data_exports") if isinstance(config.get("data_exports"), dict) else {}
    if data_exports.get("include_completed_csv"):
        files.append(_write_csv(output_dir / "cyberlab_completed_cases.csv", scoped_rows, CASE_COLUMNS))
    if data_exports.get("include_in_progress_csv"):
        files.append(_write_csv(output_dir / "cyberlab_in_progress_cases.csv", progress_rows, IN_PROGRESS_COLUMNS))
    if data_exports.get("include_summary_json"):
        files.append(_write_json(output_dir / "cyberlab_summary.json", _summary(scoped_rows, progress_rows, config)))

    graph_settings = config.get("graph_settings") if isinstance(config.get("graph_settings"), dict) else {}
    if graph_settings.get("include_csv", True) or graph_settings.get("include_png", True):
        graph_dir = _report_dir(output_dir, config, "graphs_snapshot")
        graph_dir.mkdir(parents=True, exist_ok=True)
        _remove_generated_files(graph_dir, ("cyberlab_graph_*.png", "cyberlab_graph_*.csv"))
        for graph_type in graph_settings.get("types") or ["Offense Type", "Device Type", "Agency"]:
            graph_rows = _graph_rows(graph_type, completed_rows)
            if not graph_rows:
                continue
            safe_name = _safe_filename(graph_type)
            if graph_settings.get("include_csv", True):
                files.append(_write_graph_csv(graph_dir, safe_name, graph_rows))
            if graph_settings.get("include_png", True):
                graph_png = _write_graph_png(graph_dir, graph_type, safe_name, graph_rows)
                if graph_png:
                    files.append(graph_png)

    map_settings = config.get("map_settings") if isinstance(config.get("map_settings"), dict) else {}
    if map_settings.get("include_data_file", False) or "map_html" in report_types:
        map_dir = _report_dir(output_dir, config, "map_html")
        map_dir.mkdir(parents=True, exist_ok=True)
        _remove_generated_files(map_dir, ("areas_served_map*.html", "cyberlab_map_data*.json"))
        map_points = _map_points(completed_rows, progress_rows, map_settings)
        if map_settings.get("include_data_file", False):
            files.append(_write_json(map_dir / "cyberlab_map_data.json", map_points))
        files.append(_write_map_html(map_dir / "areas_served_map.html", map_points, map_settings))

    report_cleanup_patterns = (
        "total_case_summary*.pdf",
        "date_scope_cases_summary*.pdf",
        "all_cases_summary*.pdf",
        "total_case_summary*.xlsx",
    )
    for report_type in ("reports", "total_summary_pdf", "total_summary_pdf_scope", "all_cases_pdf", "total_summary_xlsx"):
        reports_dir = _report_dir(output_dir, config, report_type)
        reports_dir.mkdir(parents=True, exist_ok=True)
        _remove_generated_files(reports_dir, report_cleanup_patterns)
    if "total_summary_pdf" in report_types:
        pdf_file = _write_summary_pdf(
            _target_file(output_dir, config, "total_summary_pdf", "total_case_summary.pdf"),
            completed_rows,
            progress_rows,
            _report_config(config, title="Total Case Summary", scope_label="All Time", detail_limit=30),
        )
        if pdf_file:
            files.append(pdf_file)
    if "total_summary_pdf_scope" in report_types:
        scope_name = _safe_filename(_scope_label(config))
        pdf_file = _write_summary_pdf(
            _target_file(output_dir, config, "total_summary_pdf_scope", f"date_scope_cases_summary_{scope_name}.pdf"),
            scoped_rows,
            progress_rows,
            _report_config(config, title="Date Scope Case Summary", detail_limit=30),
        )
        if pdf_file:
            files.append(pdf_file)
    if "all_cases_pdf" in report_types:
        pdf_file = _write_summary_pdf(
            _target_file(output_dir, config, "all_cases_pdf", "all_cases_summary.pdf"),
            completed_rows,
            progress_rows,
            _report_config(config, title="All Cases Summary", scope_label="All Time", detail_limit=max(len(completed_rows), 30)),
        )
        if pdf_file:
            files.append(pdf_file)
    if "total_summary_xlsx" in report_types:
        xlsx_file = _write_summary_xlsx(_target_file(output_dir, config, "total_summary_xlsx", "total_case_summary.xlsx"), completed_rows, progress_rows, _report_config(config, scope_label="All Time"))
        if xlsx_file:
            files.append(xlsx_file)

    return {
        "ok": True,
        "engine": "native",
        "output_dir": str(output_dir),
        "completed_cases": len(scoped_rows),
        "in_progress_cases": len(progress_rows),
        "requested_report_types": sorted(report_types),
        "pdf_files": [file for file in files if str(file.get("name", "")).lower().endswith(".pdf")],
        "report_dirs": {
            "reports": str(_report_dir(output_dir, config, "reports")),
            "graphs": str(_report_dir(output_dir, config, "graphs_snapshot")),
            "map": str(_report_dir(output_dir, config, "map_html")),
        },
        "files": files,
    }


def _resolve_output_dir(path: str | None) -> Path:
    if path and str(path).strip():
        return Path(str(path)).expanduser()
    return automated_reports_dir()


DEFAULT_REPORT_SUBDIRS = {
    "reports": "Reports",
    "total_summary_pdf": "Reports",
    "total_summary_pdf_scope": "Reports",
    "total_summary_xlsx": "Reports",
    "all_cases_pdf": "Reports",
    "graphs_snapshot": "Graphs",
    "map_html": "Areas Served Map",
}


def _report_dir(output_dir: Path, config: dict[str, Any], report_type: str) -> Path:
    report_dirs = config.get("report_output_dirs") if isinstance(config.get("report_output_dirs"), dict) else {}
    value = str(report_dirs.get(report_type) or "").strip()
    if not value and report_type in {"total_summary_pdf", "total_summary_pdf_scope", "total_summary_xlsx", "all_cases_pdf"}:
        value = str(report_dirs.get("reports") or "").strip()
    if not value:
        value = DEFAULT_REPORT_SUBDIRS.get(report_type, "")
    if not value:
        return output_dir
    path = Path(value).expanduser()
    return path if path.is_absolute() else output_dir / path


def _target_file(output_dir: Path, config: dict[str, Any], report_type: str, filename: str) -> Path:
    directory = _report_dir(output_dir, config, report_type)
    directory.mkdir(parents=True, exist_ok=True)
    return directory / filename


def _remove_generated_files(directory: Path, patterns: tuple[str, ...]) -> None:
    for pattern in patterns:
        for path in directory.glob(pattern):
            if path.is_file():
                try:
                    path.unlink()
                except OSError:
                    pass


def _read_rows(table: str, columns: list[str]) -> list[dict[str, Any]]:
    with connect() as conn:
        rows = conn.execute(f"SELECT {', '.join(columns)} FROM {table} ORDER BY datetime(created_at) DESC, id DESC").fetchall()
    return [dict(row) for row in rows]


def _parse_date(value: Any) -> datetime | None:
    if not value:
        return None
    text = str(value).strip()
    for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d", "%m/%d/%Y"):
        try:
            return datetime.strptime(text[:19] if fmt.endswith("%S") else text[:10], fmt)
        except ValueError:
            continue
    return None


def _scope_rows(rows: list[dict[str, Any]], config: dict[str, Any]) -> list[dict[str, Any]]:
    now = datetime.now()
    threshold: datetime | None = None
    mode = config.get("date_range_mode") or "all"
    if config.get("recent_only"):
        try:
            days = max(1, int(config.get("recent_days") or 31))
        except (TypeError, ValueError):
            days = 31
        threshold = now - timedelta(days=days)
    elif mode == "current_week":
        threshold = datetime(now.year, now.month, now.day) - timedelta(days=now.weekday())
    elif mode == "current_month":
        threshold = datetime(now.year, now.month, 1)

    if not threshold:
        return rows
    scoped = []
    for row in rows:
        date_value = _parse_date(row.get("end_date")) or _parse_date(row.get("start_date")) or _parse_date(row.get("created_at"))
        if date_value and date_value >= threshold:
            scoped.append(row)
    return scoped


def _summary(completed_rows: list[dict[str, Any]], progress_rows: list[dict[str, Any]], config: dict[str, Any]) -> dict[str, Any]:
    total_volume = sum(float(row.get("volume_size_gb") or 0) for row in completed_rows)
    return {
        "generated_at": datetime.now().isoformat(timespec="seconds"),
        "profile": _report_profile(),
        "scope": {
            "date_range_mode": config.get("date_range_mode") or "all",
            "recent_only": bool(config.get("recent_only")),
            "recent_days": config.get("recent_days"),
        },
        "completed_cases": len(completed_rows),
        "in_progress_cases": len(progress_rows),
        "average_turnaround_days": _average_turnaround_days(completed_rows),
        "total_volume_gb": round(total_volume, 2),
        "top_offenses": _top_counts(completed_rows, "offense_type"),
        "top_agencies": _top_counts(completed_rows, "agency"),
        "top_devices": _top_counts(completed_rows, "device_type"),
    }


def _average_turnaround_days(rows: list[dict[str, Any]]) -> float:
    durations = []
    for row in rows:
        started = _parse_date(row.get("start_date"))
        ended = _parse_date(row.get("end_date"))
        if started and ended and ended >= started:
            durations.append((ended - started).total_seconds() / 86400)
    return round(sum(durations) / len(durations), 1) if durations else 0.0


def _report_profile() -> dict[str, str]:
    stored = get_json_setting("app_profile", {})
    if not isinstance(stored, dict):
        stored = {}
    return {
        "organization": str(stored.get("organization") or "").strip(),
        "name": str(stored.get("name") or "").strip(),
    }


def _profile_lines(styles: Any) -> list[Any]:
    from reportlab.platypus import Paragraph

    profile = _report_profile()
    lines = []
    if profile["organization"]:
        lines.append(Paragraph(f"Agency/Organization: {html.escape(profile['organization'])}", styles["Normal"]))
    if profile["name"]:
        lines.append(Paragraph(f"Prepared by: {html.escape(profile['name'])}", styles["Normal"]))
    return lines


def _top_counts(rows: list[dict[str, Any]], column: str, limit: int | None = 10) -> list[dict[str, Any]]:
    counts = Counter((str(row.get(column) or "Unknown").strip() or "Unknown") for row in rows)
    items = counts.most_common(limit) if limit else counts.most_common()
    return [{"label": label, "value": value} for label, value in items]


def _geocache_lookup() -> dict[str, tuple[float, float]]:
    with connect() as conn:
        try:
            rows = conn.execute("SELECT location_key, latitude, longitude FROM geocache").fetchall()
        except Exception:
            return {}
    return {
        str(row["location_key"]): (float(row["latitude"]), float(row["longitude"]))
        for row in rows
        if row["latitude"] is not None and row["longitude"] is not None
    }


def _map_points(
    completed_rows: list[dict[str, Any]],
    progress_rows: list[dict[str, Any]] | None = None,
    map_settings: dict[str, Any] | None = None,
) -> list[dict[str, Any]]:
    settings = map_settings or {}
    grouped: dict[tuple[str, str, float | None, float | None], dict[str, Any]] = {}
    geocache = _geocache_lookup()
    source_rows: list[tuple[str, dict[str, Any]]] = []
    if settings.get("include_completed", True):
        source_rows.extend(("Completed", row) for row in completed_rows)
    if settings.get("include_in_progress", True):
        source_rows.extend(("In Progress", row) for row in (progress_rows or []))

    for status, row in source_rows:
        city_name = str(row.get("city_of_offense") or "").strip()
        location_name = str(row.get("location_name") or "").strip()
        city = location_name or city_name
        state = str(row.get("state_of_offense") or "").strip()
        if not city:
            continue
        location_key = f"{location_name}|{city_name}|{state}" if location_name else f"{city_name}|{state}"
        lat_lon = geocache.get(location_key)
        latitude = row.get("latitude") if row.get("latitude") is not None else (lat_lon[0] if lat_lon else None)
        longitude = row.get("longitude") if row.get("longitude") is not None else (lat_lon[1] if lat_lon else None)
        key = (city, state, latitude, longitude)
        item = grouped.setdefault(
            key,
            {
                "city": city,
                "state": state,
                "case_count": 0,
                "completed_count": 0,
                "in_progress_count": 0,
                "total_volume_gb": 0.0,
                "latitude": latitude,
                "longitude": longitude,
                "cases": [],
            },
        )
        item["case_count"] += 1
        if status == "Completed":
            item["completed_count"] += 1
        else:
            item["in_progress_count"] += 1
        item["total_volume_gb"] += float(row.get("volume_size_gb") or 0)
        if settings.get("include_case_details", True) and len(item["cases"]) < 12:
            item["cases"].append({
                "case_number": row.get("case_number") or "",
                "agency": row.get("agency") or "",
                "offense_type": row.get("offense_type") or "",
                "status": status,
            })
    return sorted(grouped.values(), key=lambda item: (-item["case_count"], item["city"]))


def _write_csv(path: Path, rows: list[dict[str, Any]], columns: list[str]) -> dict[str, Any]:
    with path.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(handle, fieldnames=columns, extrasaction="ignore")
        writer.writeheader()
        writer.writerows(rows)
    return _file_info(path)


def _write_json(path: Path, data: Any) -> dict[str, Any]:
    path.write_text(json.dumps(data, indent=2), encoding="utf-8")
    return _file_info(path)


def _graph_rows(graph_type: str, rows: list[dict[str, Any]], limit: int | None = None) -> list[dict[str, Any]]:
    column = GRAPH_COLUMN_MAP.get(graph_type)
    if not column:
        return []
    if graph_type in VOLUME_GRAPH_TYPES:
        totals: dict[str, float] = {}
        for row in rows:
            label = str(row.get(column) or "Unknown").strip() or "Unknown"
            totals[label] = totals.get(label, 0.0) + float(row.get("volume_size_gb") or 0)
        sorted_rows = sorted(totals.items(), key=lambda item: (-item[1], item[0]))
        if limit:
            sorted_rows = sorted_rows[:limit]
        return [{"label": label, "value": round(value, 2)} for label, value in sorted_rows]
    return _top_counts(rows, column, limit=limit)


def _safe_filename(value: str) -> str:
    return "".join(ch.lower() if ch.isalnum() else "_" for ch in value).strip("_")


def _report_config(
    config: dict[str, Any],
    *,
    title: str | None = None,
    scope_label: str | None = None,
    detail_limit: int | None = None,
) -> dict[str, Any]:
    value = dict(config)
    if title:
        value["report_title"] = title
    if scope_label:
        value["scope_label"] = scope_label
    if detail_limit is not None:
        value["detail_limit"] = detail_limit
    return value


def _write_graph_csv(output_dir: Path, safe_name: str, graph_rows: list[dict[str, Any]]) -> dict[str, Any]:
    path = output_dir / f"cyberlab_graph_{safe_name}.csv"
    with path.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(handle, fieldnames=["label", "value"])
        writer.writeheader()
        writer.writerows(graph_rows)
    return _file_info(path)


def _write_graph_png(output_dir: Path, graph_type: str, safe_name: str, graph_rows: list[dict[str, Any]]) -> dict[str, Any] | None:
    try:
        import matplotlib

        matplotlib.use("Agg")
        import matplotlib.pyplot as plt
    except ImportError as exc:
        raise RuntimeError(f"PDF export is unavailable because ReportLab could not be loaded: {exc}") from exc

    rows = list(reversed(graph_rows))
    labels = [str(row["label"]) for row in rows]
    values = [float(row["value"] or 0) for row in rows]
    height = max(4.8, 0.34 * len(rows) + 1.8)
    width = 11 if len(rows) > 16 else 10
    fig, ax = plt.subplots(figsize=(width, height), dpi=140)
    fig.patch.set_facecolor("#f6fbff")
    ax.set_facecolor("#ffffff")
    bars = ax.barh(labels, values, color="#2f82d8", edgecolor="#1d5f9f", linewidth=0.6)
    ax.set_title(graph_type, fontsize=15, fontweight="bold", color="#172033", pad=12)
    ax.set_xlabel("Total Volume" if graph_type in VOLUME_GRAPH_TYPES else "Cases", color="#44576c")
    ax.tick_params(axis="x", colors="#44576c")
    ax.tick_params(axis="y", labelsize=9, colors="#172033")
    ax.grid(axis="x", color="#d8e5f0", linewidth=0.8)
    ax.set_axisbelow(True)
    for spine in ax.spines.values():
        spine.set_visible(False)
    for bar, value in zip(bars, values):
        ax.text(
            bar.get_width(),
            bar.get_y() + bar.get_height() / 2,
            f" {_format_volume(value) if graph_type in VOLUME_GRAPH_TYPES else f'{value:g}'}",
            va="center",
            ha="left",
            fontsize=8,
            color="#172033",
        )
    fig.tight_layout()
    path = output_dir / f"cyberlab_graph_{safe_name}.png"
    fig.savefig(path, bbox_inches="tight")
    plt.close(fig)
    return _file_info(path)


def _write_map_html(path: Path, points: list[dict[str, Any]], map_settings: dict[str, Any] | None = None) -> dict[str, Any]:
    settings = map_settings or {}
    geocoded = [point for point in points if point.get("latitude") is not None and point.get("longitude") is not None]
    missing = [point for point in points if point.get("latitude") is None or point.get("longitude") is None]
    center = [geocoded[0]["latitude"], geocoded[0]["longitude"]] if geocoded else [32.7, -89.5]
    marker_data_url = ""
    if marker_icon_path().exists():
        marker_data_url = f"data:image/png;base64,{base64.b64encode(marker_icon_path().read_bytes()).decode('ascii')}"
    data_json = json.dumps({
        "points": points,
        "center": center,
        "includeCaseDetails": settings.get("include_case_details", True),
        "markerIcon": marker_data_url,
    }).replace("</", "<\\/")
    rows = "\n".join(
        "<tr>"
        f"<td>{html.escape(str(point['city']))}</td>"
        f"<td>{html.escape(str(point['state']))}</td>"
        f"<td>{point['case_count']}</td>"
        f"<td>{point['completed_count']}</td>"
        f"<td>{point['in_progress_count']}</td>"
        f"<td>{html.escape(_format_volume(point['total_volume_gb']))}</td>"
        f"<td>{'Mapped' if point.get('latitude') is not None else 'No geocode'}</td>"
        "</tr>"
        for point in points
    )
    path.write_text(
        f"""<!doctype html>
<html lang="en">
<head>
  <meta charset="utf-8">
  <meta name="viewport" content="width=device-width, initial-scale=1">
  <title>CyberLab Case Map</title>
  <link rel="stylesheet" href="https://unpkg.com/leaflet@1.9.4/dist/leaflet.css">
  <style>
    :root {{ font-family: Inter, Segoe UI, Arial, sans-serif; color: #e6f2ff; background: #06101c; }}
    body {{ margin: 0; background: linear-gradient(180deg, #0a1726, #06101c); }}
    header {{ display: flex; justify-content: space-between; align-items: center; gap: 16px; padding: 18px 22px; border-bottom: 1px solid #2e5e86; background: #0a2035; }}
    h1 {{ margin: 0; font-size: 21px; }}
    .muted {{ color: #9ab5cd; }}
    .metrics {{ display: flex; gap: 10px; flex-wrap: wrap; }}
    .metric {{ padding: 8px 12px; border: 1px solid #2e5e86; border-radius: 999px; background: #081a2c; }}
    #map {{ height: min(72vh, 760px); min-height: 520px; }}
    main {{ padding: 18px; }}
    table {{ border-collapse: collapse; width: 100%; margin-top: 18px; overflow: hidden; border-radius: 14px; }}
    th, td {{ border-bottom: 1px solid #254b69; padding: 9px 11px; text-align: left; }}
    th {{ background: #0a2035; color: #e6f2ff; }}
    td {{ color: #d6e7f7; background: #081a2c; }}
    .marker-pin {{ display: grid; place-items: center; width: 34px; height: 34px; border: 2px solid #e6f2ff; border-radius: 50%; color: #06101c; background: #62b9ff; font-weight: 800; box-shadow: 0 6px 18px rgba(0,0,0,.32); }}
    .custom-marker-icon {{ border: 0 !important; background: transparent !important; object-fit: contain; filter: drop-shadow(0 7px 10px rgba(0,0,0,.35)); }}
    .popup-title {{ font-weight: 800; color: #172033; }}
    .case-list {{ margin: 8px 0 0; padding-left: 16px; }}
    .missing {{ margin-top: 10px; color: #ffcf75; }}
  </style>
</head>
<body>
  <header>
    <div>
      <h1>CyberLab Case Map</h1>
      <div class="muted">Generated {html.escape(datetime.now().strftime("%Y-%m-%d %H:%M:%S"))}</div>
    </div>
    <div class="metrics">
      <div class="metric">{len(points)} locations</div>
      <div class="metric">{sum(int(point["case_count"]) for point in points)} cases</div>
      <div class="metric">{len(geocoded)} mapped</div>
    </div>
  </header>
  <div id="map"></div>
  <main>
    <div class="muted">Locations without geocache coordinates are listed in the table and can be mapped after the app geocache is updated.</div>
    <table>
      <thead><tr><th>City</th><th>State</th><th>Total Cases</th><th>Completed</th><th>In Progress</th><th>Total Volume</th><th>Status</th></tr></thead>
      <tbody>{rows}</tbody>
    </table>
  </main>
  <script src="https://unpkg.com/leaflet@1.9.4/dist/leaflet.js"></script>
  <script>
    const exportData = {data_json};
    const map = L.map('map', {{ scrollWheelZoom: true }}).setView(exportData.center, exportData.points.length ? 7 : 5);
    L.tileLayer('https://{{s}}.tile.openstreetmap.org/{{z}}/{{x}}/{{y}}.png', {{
      maxZoom: 18,
      attribution: '&copy; OpenStreetMap contributors'
    }}).addTo(map);
    const bounds = [];
    const markerSizeForZoom = (zoom) => {{
      const level = Number(zoom || 6);
      if (level <= 4) return 8;
      if (level <= 5) return 10;
      if (level <= 6) return 13;
      if (level <= 7) return 17;
      if (level <= 8) return 22;
      return Math.min(44, 22 + (level - 8) * 5);
    }};
    const formatVolume = (gbValue) => {{
      const gb = Number(gbValue || 0);
      if (!Number.isFinite(gb)) return '-';
      if (Math.abs(gb) >= 1024) return `${{(gb / 1024).toFixed(1)}} TB`;
      return `${{gb.toFixed(1)}} GB`;
    }};
    const activeMarkers = [];
    function buildIcon(point) {{
      const size = markerSizeForZoom(map.getZoom());
      if (exportData.markerIcon) {{
        return L.icon({{
          iconUrl: exportData.markerIcon,
          iconSize: [size, size],
          iconAnchor: [size / 2, size / 2],
          popupAnchor: [0, -size / 2],
          className: 'custom-marker-icon',
        }});
      }}
      return L.divIcon({{
        className: '',
        html: `<span class="marker-pin" style="width:${{size}}px;height:${{size}}px;font-size:${{Math.max(9, size * 0.32)}}px;">${{point.case_count}}</span>`,
        iconSize: [size, size],
        iconAnchor: [size / 2, size / 2],
        popupAnchor: [0, -size / 2],
      }});
    }}
    exportData.points
      .filter((point) => point.latitude !== null && point.longitude !== null)
      .forEach((point) => {{
        const latLng = [Number(point.latitude), Number(point.longitude)];
        bounds.push(latLng);
        const icon = buildIcon(point);
        const cases = exportData.includeCaseDetails && point.cases?.length
          ? `<ul class="case-list">${{point.cases.map((item) => `<li>${{item.case_number || 'No case #'}} - ${{item.status}}</li>`).join('')}}</ul>`
          : '';
        const marker = L.marker(latLng, {{ icon }}).addTo(map).bindPopup(`
          <div class="popup-title">${{point.city}}, ${{point.state}}</div>
          <div>${{point.case_count}} cases / ${{formatVolume(point.total_volume_gb)}}</div>
          <div>${{point.completed_count}} completed / ${{point.in_progress_count}} in progress</div>
          ${{cases}}
        `);
        activeMarkers.push([marker, point]);
      }});
    map.on('zoomend', () => activeMarkers.forEach(([marker, point]) => marker.setIcon(buildIcon(point))));
    if (bounds.length > 1) {{
      map.fitBounds(bounds, {{ padding: [38, 38] }});
    }}
    if ({len(missing)} > 0) {{
      const missingControl = L.control({{ position: 'bottomleft' }});
      missingControl.onAdd = function() {{
        const div = L.DomUtil.create('div', 'missing');
        div.innerHTML = '{len(missing)} locations need geocoding';
        return div;
      }};
      missingControl.addTo(map);
    }}
  </script>
</body>
</html>
""",
        encoding="utf-8",
    )
    return _file_info(path)


def _write_summary_pdf(path: Path, completed_rows: list[dict[str, Any]], progress_rows: list[dict[str, Any]], config: dict[str, Any]) -> dict[str, Any] | None:
    try:
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import A4, landscape, legal, letter, portrait
        from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
        from reportlab.lib.units import inch
        from reportlab.platypus import Image, KeepTogether, PageBreak, Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle
    except ImportError:
        return None

    base_size = {"Letter": letter, "Legal": legal, "A4": A4}.get(config.get("page_size"), letter)
    orientation = config.get("orientation") or "Auto"
    use_landscape = orientation == "Landscape" or (orientation == "Auto" and len(completed_rows) > 80)
    page_size = landscape(base_size) if use_landscape else portrait(base_size)
    doc = SimpleDocTemplate(str(path), pagesize=page_size, leftMargin=24, rightMargin=24, topMargin=28, bottomMargin=28)
    styles = getSampleStyleSheet()
    styles.add(ParagraphStyle("Small", parent=styles["Normal"], fontSize=8, leading=10))
    styles.add(ParagraphStyle("Tiny", parent=styles["Normal"], fontSize=7, leading=8))
    summary = _summary(completed_rows, progress_rows, config)
    page_width = page_size[0] - doc.leftMargin - doc.rightMargin
    elements = []
    report_title = str(config.get("report_title") or "CyberLab Case Summary")
    title_block: Any = Paragraph(report_title, styles["Title"])
    if logo_path().exists():
        try:
            logo = Image(str(logo_path()), width=0.85 * inch, height=0.85 * inch)
            title_block = Table([[Paragraph(report_title, styles["Title"]), logo]], colWidths=[page_width - inch, inch])
        except Exception:
            title_block = Paragraph(report_title, styles["Title"])
    elements.extend([
        title_block,
        Paragraph(f"Generated: {summary['generated_at']} / Scope: {_scope_label(config)}", styles["Normal"]),
        *_profile_lines(styles),
        Spacer(1, 12),
        _styled_table(
            [
                ["Completed Cases", "In-Progress Cases", "Avg. Turnaround", "Total Volume", "Mapped Locations"],
                [
                    str(summary["completed_cases"]),
                    str(summary["in_progress_cases"]),
                    f"{summary['average_turnaround_days']:.1f} days",
                    _format_volume(summary["total_volume_gb"]),
                    str(len(_map_points(completed_rows, progress_rows, {"include_completed": True, "include_in_progress": True, "include_case_details": False}))),
                ],
            ],
            col_widths=[page_width * 0.2] * 5,
            colors=colors,
        ),
        Spacer(1, 12),
    ])

    breakdowns = [
        ("Top Offenses", "offense_type"),
        ("Top Agencies", "agency"),
        ("Device Types", "device_type"),
        ("Forensic Tools", "forensic_tool"),
        ("Examiners", "examiner"),
    ]
    for title, column in breakdowns:
        rows = _top_counts(completed_rows, column, limit=12)
        if not rows:
            continue
        elements.append(Paragraph(title, styles["Heading2"]))
        elements.append(_styled_table([["Value", "Count"], *[[item["label"], item["value"]] for item in rows]], [page_width - 1.1 * inch, 1.1 * inch], colors))
        elements.append(Spacer(1, 8))

    volume_rows = _graph_rows("Total Volume by Agency", completed_rows, limit=10)
    if volume_rows:
        elements.append(KeepTogether([
            Paragraph("Top Agencies by Volume", styles["Heading2"]),
            _styled_table([["Agency", "Volume"], *[[item["label"], _format_volume(item["value"])] for item in volume_rows]], [page_width - 1.4 * inch, 1.4 * inch], colors),
            Spacer(1, 8),
        ]))

    aging_rows = _case_aging_rows(progress_rows)
    if aging_rows:
        elements.append(PageBreak())
        elements.append(Paragraph("Case Aging Alerts", styles["Heading2"]))
        elements.append(_styled_table([["Case #", "Agency", "Priority", "Due", "Status", "Days"], *aging_rows], _fit_widths(page_width, [1.2, 1.8, 0.9, 1.0, 1.0, 0.6]), colors, font_size=8))
        elements.append(Spacer(1, 10))

    try:
        detail_limit = max(1, int(config.get("detail_limit") or 30))
    except (TypeError, ValueError):
        detail_limit = 30
    recent_cases = completed_rows[: min(detail_limit, len(completed_rows))]
    if recent_cases:
        elements.append(PageBreak())
        case_section_title = "Completed Cases" if detail_limit >= len(completed_rows) else "Recent Completed Cases"
        elements.append(Paragraph(case_section_title, styles["Heading2"]))
        recent_table = [["Case #", "Created", "Examiner", "Agency", "Offense", "Device", "Volume"]]
        for row in recent_cases:
            recent_table.append([
                row.get("case_number") or "",
                _short_date(row.get("created_at") or row.get("start_date")),
                row.get("examiner") or "",
                Paragraph(html.escape(str(row.get("agency") or "")), styles["Tiny"]),
                Paragraph(html.escape(str(row.get("offense_type") or "")), styles["Tiny"]),
                row.get("device_type") or "",
                _format_volume(row.get("volume_size_gb")),
            ])
        elements.append(_styled_table(recent_table, _fit_widths(page_width, [1.1, 0.9, 1.1, 1.35, 1.35, 1.1, 0.55]), colors, font_size=7.5))

    if progress_rows:
        elements.append(Spacer(1, 12))
        elements.append(Paragraph("In-Progress Cases", styles["Heading2"]))
        progress_table = [["Case #", "Agency", "Offense", "Priority", "Workflow", "Due"]]
        for row in progress_rows[:30]:
            progress_table.append([
                row.get("case_number") or "",
                row.get("agency") or "",
                row.get("offense_type") or "",
                row.get("priority") or "",
                row.get("workflow_status") or "",
                _short_date(row.get("target_due_date")),
            ])
        elements.append(_styled_table(progress_table, _fit_widths(page_width, [1.1, 1.5, 1.5, 0.8, 1.2, 0.9]), colors, font_size=7.5))

    try:
        doc.build(elements)
    except Exception as exc:
        raise RuntimeError(f"Failed to write PDF report '{path}': {exc}") from exc
    return _file_info(path)


def _scope_label(config: dict[str, Any]) -> str:
    if config.get("scope_label"):
        return str(config["scope_label"])
    if config.get("recent_only"):
        return f"Last {config.get('recent_days') or 31} days"
    return str(config.get("date_range_mode") or "all").replace("_", " ").title()


def _format_volume(value: Any) -> str:
    gb = float(value or 0)
    if gb >= 1024:
        return f"{gb / 1024:.2f} TB"
    return f"{gb:.2f} GB"


def _short_date(value: Any) -> str:
    parsed = _parse_date(value)
    return parsed.strftime("%Y-%m-%d") if parsed else (str(value or "")[:10])


def _fit_widths(page_width: float, weights: list[float]) -> list[float]:
    total = sum(weights) or 1
    return [(weight / total) * page_width for weight in weights]


def _styled_table(
    rows: list[list[Any]],
    col_widths: list[float],
    colors: Any,
    font_size: float = 8.5,
) -> Any:
    from reportlab.platypus import Table, TableStyle

    table = Table(rows, colWidths=col_widths, repeatRows=1, hAlign="LEFT")
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#dbeafe")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.HexColor("#172033")),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("GRID", (0, 0), (-1, -1), 0.45, colors.HexColor("#9db8d1")),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f6fbff")]),
        ("PADDING", (0, 0), (-1, -1), 5),
        ("FONTSIZE", (0, 0), (-1, -1), font_size),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
    ]))
    return table


def _case_aging_rows(progress_rows: list[dict[str, Any]]) -> list[list[Any]]:
    today = datetime.now().date()
    rows = []
    for row in progress_rows:
        due = _parse_date(row.get("target_due_date"))
        if not due:
            continue
        delta = (due.date() - today).days
        if delta < 0:
            status = "Overdue"
            days = abs(delta)
        elif delta == 0:
            status = "Due Today"
            days = 0
        elif delta <= 7:
            status = "Due Soon"
            days = delta
        else:
            continue
        rows.append([
            row.get("case_number") or "",
            row.get("agency") or "",
            row.get("priority") or "",
            _short_date(row.get("target_due_date")),
            status,
            str(days),
        ])
    return sorted(rows, key=lambda item: (item[4] != "Overdue", item[3]))


def _write_summary_xlsx(path: Path, completed_rows: list[dict[str, Any]], progress_rows: list[dict[str, Any]], config: dict[str, Any]) -> dict[str, Any] | None:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter
    except ImportError:
        return None

    wb = Workbook()
    summary = _summary(completed_rows, progress_rows, config)

    ws = wb.active
    ws.title = "Summary"
    _write_title(ws, "CyberLab Case Summary")
    profile = _report_profile()
    summary_rows = [
        ("Generated", summary["generated_at"]),
        ("Agency/Organization", profile["organization"] or "-"),
        ("Prepared By", profile["name"] or "-"),
        ("Scope", _scope_label(config)),
        ("Completed Cases", summary["completed_cases"]),
        ("In-Progress Cases", summary["in_progress_cases"]),
        ("Average Turnaround", f"{summary['average_turnaround_days']:.1f} days"),
        ("Total Volume", _format_volume(summary["total_volume_gb"])),
        ("Mapped Locations", len(_map_points(completed_rows, progress_rows, {"include_completed": True, "include_in_progress": True, "include_case_details": False}))),
    ]
    for row in summary_rows:
        ws.append(row)
    _style_sheet(ws)

    for title, column in [
        ("Top Offenses", "offense_type"),
        ("Top Agencies", "agency"),
        ("Device Types", "device_type"),
        ("Forensic Tools", "forensic_tool"),
        ("Examiners", "examiner"),
    ]:
        sheet = wb.create_sheet(_sheet_title(title))
        _write_title(sheet, title)
        sheet.append(["Value", "Count"])
        for item in _top_counts(completed_rows, column, limit=200):
            sheet.append([item["label"], item["value"]])
        _style_sheet(sheet)

    volume_sheet = wb.create_sheet("Volume by Agency")
    _write_title(volume_sheet, "Top Agencies by Volume")
    volume_sheet.append(["Agency", "Volume GB", "Formatted"])
    for item in _graph_rows("Total Volume by Agency", completed_rows, limit=200):
        volume_sheet.append([item["label"], item["value"], _format_volume(item["value"])])
    _style_sheet(volume_sheet)

    cases_sheet = wb.create_sheet("Completed Cases")
    _write_title(cases_sheet, "Completed Cases")
    cases_sheet.append(CASE_COLUMNS)
    for row in completed_rows:
        cases_sheet.append([row.get(column) for column in CASE_COLUMNS])
    _style_sheet(cases_sheet, freeze_row=3)

    progress_sheet = wb.create_sheet("In Progress")
    _write_title(progress_sheet, "In-Progress Cases")
    progress_sheet.append(IN_PROGRESS_COLUMNS)
    for row in progress_rows:
        progress_sheet.append([row.get(column) for column in IN_PROGRESS_COLUMNS])
    _style_sheet(progress_sheet, freeze_row=3)

    aging_rows = _case_aging_rows(progress_rows)
    aging_sheet = wb.create_sheet("Aging Alerts")
    _write_title(aging_sheet, "Case Aging Alerts")
    aging_sheet.append(["Case #", "Agency", "Priority", "Due", "Status", "Days"])
    for row in aging_rows:
        aging_sheet.append(row)
    _style_sheet(aging_sheet, freeze_row=3)

    map_sheet = wb.create_sheet("Map Locations")
    _write_title(map_sheet, "Map Locations")
    map_sheet.append(["City", "State", "Cases", "Completed", "In Progress", "Total Volume", "Latitude", "Longitude"])
    for point in _map_points(completed_rows, progress_rows, {"include_completed": True, "include_in_progress": True, "include_case_details": False}):
        map_sheet.append([
            point["city"],
            point["state"],
            point["case_count"],
            point["completed_count"],
            point["in_progress_count"],
            _format_volume(point["total_volume_gb"]),
            point["latitude"],
            point["longitude"],
        ])
    _style_sheet(map_sheet)

    wb.save(path)
    return _file_info(path)


def _sheet_title(title: str) -> str:
    invalid = set("[]:*?/\\")
    cleaned = "".join("_" if ch in invalid else ch for ch in title)[:31]
    return cleaned or "Sheet"


def _write_title(ws: Any, title: str) -> None:
    from openpyxl.styles import Font, PatternFill

    ws["A1"] = title
    ws["A1"].font = Font(bold=True, size=14, color="172033")
    ws["A1"].fill = PatternFill("solid", fgColor="DBEAFE")
    ws.freeze_panes = "A3"


def _style_sheet(ws: Any, freeze_row: int = 2) -> None:
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    header_fill = PatternFill("solid", fgColor="DBEAFE")
    header_font = Font(bold=True, color="172033")
    ws.freeze_panes = f"A{freeze_row}"
    for cell in ws[2]:
        if cell.value is not None:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")
    for column in ws.columns:
        max_len = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            value = "" if cell.value is None else str(cell.value)
            max_len = max(max_len, min(len(value), 60))
            cell.alignment = Alignment(vertical="top", wrap_text=True)
        ws.column_dimensions[column_letter].width = max(11, min(max_len + 2, 42))


def _file_info(path: Path) -> dict[str, Any]:
    stat = path.stat()
    return {
        "name": path.name,
        "path": str(path),
        "size": stat.st_size,
        "modified": stat.st_mtime,
    }
